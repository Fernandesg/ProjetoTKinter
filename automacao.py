from tkinter.filedialog import askopenfilenames
from tkinter.ttk import *
from tkinter import *
from tkcalendar import DateEntry 
from mttkinter import mtTkinter
from time import sleep
from playwright.sync_api import sync_playwright, TimeoutError
from datetime import datetime, date, timedelta
import smtplib
from openpyxl import Workbook, load_workbook
import subprocess
from win10toast import ToastNotifier
from threading import *

credencialEmail = open('credencialEmail_AUT.txt', 'r')
loginEmail = []

for linhas in credencialEmail:
    linhas = linhas.strip()
    loginEmail.append(linhas)
usuario_email = loginEmail[0][17:-1]
senha_email = loginEmail[1][15:-1]
s = smtplib.SMTP('smtp.gmail.com: 587')
s.starttls()
s.login(usuario_email, senha_email)

dicioLogin = {}
with open('credenciais.txt', 'r') as passwords2 :
    for linha in passwords2:
        (x,y) = linha.split('=')
        x = x.strip()
        y= y.strip()
        dicioLogin[x]=y
dicioLogin['site']

with open('filiais.txt', 'r', encoding="UTF-8") as filiais_caminho :
    filiais = []
    for linhas in filiais_caminho:
        linhas = linhas.strip()
        filiais.append(linhas)

listaTipo = []
dicioTipo = {}
with open("TipoRequisicao.txt", encoding="ISO-8859-1") as dicionarioTipoReq:
    for line in dicionarioTipoReq:
        if line != '':
            (k, v) = line.split(';', 1)
            dicioTipo[str(k)] = v.strip()
    for chave in dicioTipo.keys():
        if chave != '':
            listaTipo.append(chave)

with open('codigos.txt', 'r', encoding="UTF-8") as cod_caminho :
    codigos = []
    for linhas in cod_caminho:
        linhas = linhas.strip()
        if linhas != '':
            codigos.append(linhas)

with open('centrocustos.txt', 'r', encoding="UTF-8") as cc_caminho:
    centroCustos = []
    for linhas in cc_caminho:
        linhas = linhas.strip()
        if linhas != '':
            centroCustos.append(linhas)

with open('categorias.txt', 'r', encoding="UTF-8") as categorias_caminho :
    categorias = []
    for linhas in categorias_caminho:
        linhas = linhas.strip()
        if linhas != '':
            categorias.append(linhas)

dicioConfig = {}
with open('configExcel.txt', 'r', encoding='ISO-8859-1') as configexcel:
    for item in configexcel:
        x,y = item.split(';')
        dicioConfig[x]=y.strip()

tempoMonitor = int(dicioConfig['TEMPO'])
nomeplanilha = dicioConfig['NOMEARQUIVO'].split('/')[-1]
nomeplanilhaSemext=nomeplanilha.split('.')

janela = Tk()
janela.title('Abertura de requisições')
janela.geometry('350x600')

menubar = Menu(janela)
arquivomenu = Menu(menubar, tearoff=0)
menubar.add_cascade(label="Arquivos", menu=arquivomenu)
arquivomenu.add_command(label='Itens', command=lambda: abreArquivo('codigos.txt'))
arquivomenu.add_command(label='Categorias', command=lambda: abreArquivo('categorias.txt'))
arquivomenu.add_command(label='Centro de custos', command=lambda: abreArquivo('centrocustos.txt'))
arquivomenu.add_command(label='Tipo requisição', command=lambda: abreArquivo('TipoRequisicao.txt'))
arquivomenu.add_command(label='Filiais', command=lambda: abreArquivo('filiais.txt'))
arquivomenu.add_separator()
arquivomenu.add_command(label='Credenciais ME', command=lambda: abreArquivo('credenciais.txt'))

configmenu = Menu(menubar, tearoff=0)
menubar.add_cascade(label="Configurações", menu=configmenu)
configmenu.add_command(label='Configurar monitor', command=lambda: configMonitor())

toaster = ToastNotifier()
codLista = []


def configMonitor():
    global EntryNF 
    global entryRequisicao
    global entryDataAbertura 
    global entryCNPJ 
    global entryNumPrePedido
    global entryDataPrePedido 
    global entryPedido 
    global entryStatus 
    global entryNomeAba 
    global janelaConfig
    global entryNomeArquivo
    global entryTempoMonitor

    janelaConfig = Toplevel(janela)
    janelaConfig.title(f'Editar configurações')
    janelaConfig.geometry('400x300')

    EntryNF = StringVar()
    entryRequisicao = StringVar()
    entryDataAbertura = StringVar()
    entryCNPJ = StringVar()
    entryNumPrePedido = StringVar()
    entryDataPrePedido = StringVar()
    entryPedido = StringVar()
    entryStatus = StringVar()
    entryNomeAba = StringVar()
    entryNomeArquivo = StringVar()
    entryTempoMonitor = StringVar()

    labelTitulo = Label(janelaConfig,text='Inserir as letras das colunas no Excel!', font='calibri 11')
    labelTitulo.place(relx=.03, rely=.05)

    labelTempoMonitor = Label(janelaConfig,text='Tempo monitor (min)', font='calibri 11')
    labelTempoMonitor.place(relx=.65, rely=.01)
    entryTempoMonitor = Entry(janelaConfig, width=10, textvariable=entryTempoMonitor)
    entryTempoMonitor.insert(0,dicioConfig['TEMPO'])
    entryTempoMonitor.place(relx=0.75, rely = .08)

    labelEntryNF = Label(janelaConfig,text='NF', font='calibri 10')
    labelEntryNF.place(relx=.03, rely=.2)
    entryNF = Entry(janelaConfig, width=10, textvariable=EntryNF)
    entryNF.insert(0,dicioConfig['NF'])
    entryNF.place(relx=.03, rely = .3)

    labelentryRequisicao = Label(janelaConfig,text='Requisição', font='calibri 10')
    labelentryRequisicao.place(relx=.26, rely=.2)
    entryRequisicao = Entry(janelaConfig, width=10, textvariable=entryRequisicao)
    entryRequisicao.insert(0,dicioConfig['REQUISICAO'])
    entryRequisicao.place(relx=.26, rely = .3)

    labelentryDataAbertura = Label(janelaConfig,text='Data de abertura', font='calibri 10')
    labelentryDataAbertura.place(relx=.5, rely=.2)
    entryDataAbertura = Entry(janelaConfig, width=10, textvariable=entryDataAbertura)
    entryDataAbertura.insert(0,dicioConfig['DATAABERTURA'])
    entryDataAbertura.place(relx=.5, rely = .3)

    labelentryCNPJ = Label(janelaConfig,text='CNPJ fornecedor', font='calibri 10')
    labelentryCNPJ.place(relx=.75, rely=.2)
    entryCNPJ = Entry(janelaConfig, width=10, textvariable=entryCNPJ)
    entryCNPJ.insert(0,dicioConfig['CNPJ'])
    entryCNPJ.place(relx=0.75, rely = .3)

    labelentryNumPrePedido = Label(janelaConfig,text='Nº pré-pedido', font='calibri 10')
    labelentryNumPrePedido.place(relx=.03, rely=.4)
    entryNumPrePedido = Entry(janelaConfig, width=10, textvariable=entryNumPrePedido)
    entryNumPrePedido.insert(0,dicioConfig['NUMPREPEDIDO'])
    entryNumPrePedido.place(relx=0.03, rely = .5)

    labelentryDataPrePedido = Label(janelaConfig,text='Data Pré-Pedido', font='calibri 10')
    labelentryDataPrePedido.place(relx=.26, rely=.4)
    entryDataPrePedido = Entry(janelaConfig, width=10, textvariable=entryDataPrePedido)
    entryDataPrePedido.insert(0,dicioConfig['DATAPREPEDIDO'])
    entryDataPrePedido.place(relx=.26, rely = .5)

    labelentryPedido = Label(janelaConfig,text='Nº Pedido', font='calibri 10')
    labelentryPedido.place(relx=.5, rely=.4)
    entryPedido = Entry(janelaConfig, width=10, textvariable=entryPedido)
    entryPedido.insert(0,dicioConfig['PEDIDO'])
    entryPedido.place(relx=0.5, rely = .5)

    labelentryStatus = Label(janelaConfig,text='Status', font='calibri 10')
    labelentryStatus.place(relx=.75, rely=.4)
    entryStatus = Entry(janelaConfig, width=10, textvariable=entryStatus)
    entryStatus.insert(0,dicioConfig['STATUS'])
    entryStatus.place(relx=.75, rely = .5)
    
    labelentryNomeAba = Label(janelaConfig,text='Nome da aba', font='calibri 10')
    labelentryNomeAba.place(relx=.03, rely=.6)
    entryNomeAba = Entry(janelaConfig, width=25, textvariable=entryNomeAba)
    entryNomeAba.insert(0,dicioConfig['NOMEABA'])
    entryNomeAba.place(relx=.03, rely = .7)

    labelentryNomeArquivo = Label(janelaConfig,text='Local do arquivo', font='calibri 10')
    labelentryNomeArquivo.place(relx=.5, rely=.6)
    entryNomeArquivo = Entry(janelaConfig, width=20, textvariable=entryNomeArquivo)
    entryNomeArquivo.insert(0,dicioConfig['NOMEARQUIVO'])
    entryNomeArquivo.place(relx=.5, rely = .7)
    botaoProcuraArquivo = Button(janelaConfig, text='Procurar', command=lambda: procurarArquivos('configExcel'), width=6, font='calibri, 10')
    botaoProcuraArquivo.place(relx=.82, rely=.68)

    botaoSalvar = Button(janelaConfig,text='Salvar', font='calibri 11', width=8, command=lambda: salvarArquivo('configExcel.txt'))
    botaoSalvar.place(relx=.03, rely=.8)
    botaoCancelar = Button(janelaConfig,text='Cancelar', font='calibri 11' ,width=8, command=janelaConfig.destroy)
    botaoCancelar.place(relx=.75, rely = .8)

def monitorME():
    tabela = load_workbook(dicioConfig['NOMEARQUIVO'], data_only=True)
    aba_ativa = tabela[dicioConfig['NOMEABA']]
    sleep(15)
    global varmonitorReq
    while True:

        quantPendente = 0
        if varmonitorReq.get():
            for celula in aba_ativa[dicioConfig['PEDIDO']]:
                linha = celula.row
             
                if celula.value == None and aba_ativa[dicioConfig['STATUS']+str(linha)].value == 'Pendente':
                    quantPendente += 1
            
            toaster.show_toast(f'Você possui {quantPendente} requisições pendentes!',f'O sistema irá monitorar as requisições a cada {tempoMonitor} minuto (s).',icon_path='iconeME.ico', duration=10, threaded=True)
            
            if quantPendente > 0:
                
                try:
                    with sync_playwright() as p:
                        browser = p.chromium.launch(channel="chrome")
                        page = browser.new_page()
                        page.goto(dicioLogin['site'])

                        # LOGIN ME
                        page.locator('xpath=//*[@id="LoginName"]').fill(dicioLogin['usuario_me'])
                        page.locator('xpath=//*[@id="RAWSenha"]').fill(dicioLogin['senha_me'])
                        page.locator('xpath=//*[@id="SubmitAuth"]').click()
                        page.wait_for_timeout(1)

                        # ANALISA STATUS DA REQUISIÇÃO E ATUALIZA PLANILHA
                        for celula in aba_ativa[dicioConfig['STATUS']]:
                            linha = celula.row
                            if celula.value == 'Pendente' and aba_ativa[dicioConfig['NUMPREPEDIDO']+str(linha)].value == None:
                                cnpj = str(aba_ativa[dicioConfig['CNPJ'] + str(linha)].value)
                                reqPendente = aba_ativa[dicioConfig['REQUISICAO']+str(linha)].value
                                page.goto(f'https://www.me.com.br/DO/Request/Home.mvc/Show/{reqPendente}')
                                tituloreq = page.locator('//*[@id="formRequest"]/div/div[2]/div[1]/p[1]').inner_html().strip()
                                statusRequisicao = page.locator('//*[@id="formRequest"]/div/div[2]/div[2]/p[2]/span[2]').inner_html().strip()
                                filial_requisicao = page.locator('//*[@id="formRequest"]/section[1]/div[1]/div[2]').inner_html().strip()
                                try:
                                    statusGeral = page.locator('.icon-status-item-stagethree+ b').inner_text().strip()  
                                except:
                                    statusGeral = page.locator('#formItemContext1 .green+ b').inner_text().strip()
                                statusPrePedidoTemp = statusGeral.split()[1].strip()
                                numPrePedidoTemp = statusGeral.split()[-1].strip()

                                if statusRequisicao == 'APROVADO' and statusGeral == 'Em Pendência de Compra':
                                #CRIAR PRE-PEDIDO
                                    toaster.show_toast(f'Requisição aprovada!',f'Requisição {reqPendente} aprovada! \n Criando pré-pedido',icon_path='iconeME.ico', duration=10, threaded=True)
                                    page.locator('xpath=//*[@id="btnEmergency"]').click()
                                    page.locator('xpath=/html/body/div[1]/div[3]/div/button[1]/span').click()
                                    page.locator('xpath=//*[@id="MEComponentManager_MEButton_2"]').click()
                                    page.locator('xpath=//*[@id="CGC"]').fill(cnpj)
                                    page.keyboard.press('Enter')
                                    page.locator('xpath=//*[@id="grid"]/div[2]/table/tbody/tr/td[1]/div/input').click()
                                    page.locator('xpath=//*[@id="btnSalvarSelecao"]').click()
                                    page.locator('xpath=//*[@id="btnVoltarPrePedEmergencial"]').click()
                                    page.locator('xpath=//*[@id="Resumo"]').fill(tituloreq)
                                    dataesperada = page.locator('xpath=/html/body/main/form[2]/table[1]/tbody/tr[5]/td').inner_html()
                                    
                                    if dataesperada < date.today().strftime('%d/%m/%Y'):
                                        datamais1 = timedelta(1)+date.today()
                                        page.locator('//*[@id="DataEntrega"]').fill(datamais1.strftime('%d/%m/%Y'))
                                    else: 
                                        page.locator('//*[@id="DataEntrega"]').fill(dataesperada)

                                    filiaisPrePedido = page.locator('//select[@name="LocalCobranca"]').inner_html().split('\n')
                                    for i in filiaisPrePedido:
                                        if filial_requisicao in i:
                                            indice = filiaisPrePedido.index(i)
                                            break

                                    page.locator('//select[@name="LocalCobranca"]').select_option(index=indice-1)
                                    page.locator('xpath=//*[@id="MEComponentManager_MEButton_3"]').click()
                                    page.locator('xpath=/html/body/main/form[2]/table[3]/tbody/tr[1]/td/input[1]').click()
                                    page.locator('xpath=//*[@id="MEComponentManager_MEButton_2"]').click()
                                    page.locator('xpath=//*[@id="MEComponentManager_MEButton_2"]').click()
                                    page.locator('xpath=//*[@id="formItemStatusHistory"]/div/b[1]/a').click()
                                    numPrePedido = page.locator('xpath=/html/body/main/div/div[1]/div[1]/p').inner_html().strip()
                                    statusPrePedido = page.locator('xpath=/html/body/main/div/div[1]/div[2]/div[2]/p[1]/span[2]').inner_html().strip()
                                    aba_ativa[dicioConfig['DATAPREPEDIDO'] + str(linha)] = date.today().strftime('%d/%m/%Y')
                                    aba_ativa[dicioConfig['NUMPREPEDIDO'] + str(linha)] = numPrePedido
                                    sleep(5)

                                elif statusPrePedidoTemp == 'Pré-Pedido':
                                    page.goto(f'https://www.me.com.br/VerPrePedidoWF.asp?Pedido={numPrePedidoTemp}&SuperCleanPage=false&Origin=home')
                                    statusPrePedido = page.locator('xpath=/html/body/main/div/div[1]/div[2]/div[2]/p[1]/span[2]').inner_html().strip()
                                    if statusPrePedido == 'RECUSADO':
                                        toaster.show_toast(f'Requisição {reqPendente} foi recusada!',f'O pré-pedido {numPrePedidoTemp} foi recusado, verifique o motivo.',icon_path='iconeME.ico', duration=8, threaded=True)
                                        aba_ativa[dicioConfig['PEDIDO'] + str(linha)] = statusPrePedido
                                    else: 
                                        toaster.show_toast(f'Requisição {reqPendente} já possui pré-pedido aprovado!',f'O número do pré-pedido é {numPrePedidoTemp}. \nPlanilha atualizada e monitorando!',icon_path='iconeME.ico', duration=8, threaded=True)
                                    dataaberturatemp = page.locator('xpath=//*[@id="ctl00_conteudo_gridStatus_pnlGridHistoricoV2"]/table/tbody/tr[2]/td[3]').inner_html().split()[0]
                                    aba_ativa[dicioConfig['DATAPREPEDIDO'] + str(linha)] = dataaberturatemp
                                    aba_ativa[dicioConfig['NUMPREPEDIDO'] + str(linha)] = numPrePedidoTemp
                                    sleep(5)
                                    continue

                                elif statusPrePedidoTemp == 'Pedido:':
                                    page.goto(f'https://www.me.com.br/ShowPedido.asp?Pedido={numPrePedidoTemp}&SuperCleanPage=false')
                                    numPedidoTemp = page.locator('xpath=/html/body/main/div/div[2]/div[1]/p[1]').inner_html().strip()
                                    dataaberturatemp = page.locator('xpath=//*[@id="MEComponentManager_MECollapse_2"]/div/div/div[4]/div/div/span').inner_html().split()[0]
                                    aba_ativa[dicioConfig['NUMPREPEDIDO'] + str(linha)] = numPrePedidoTemp
                                    aba_ativa[dicioConfig['PEDIDO'] + str(linha)] = numPedidoTemp
                                    aba_ativa[dicioConfig['DATAPREPEDIDO'] + str(linha)] = dataaberturatemp
                                    toaster.show_toast(f'Requisição {reqPendente} já possui pedido aprovado!',f'O número do pedido é {numPedidoTemp}. Planilha atualizada e monitorando!',icon_path='iconeME.ico', duration=8, threaded=True)
                                    sleep(5)
                                    continue

                            if celula.value == 'Pendente' and aba_ativa[dicioConfig['NUMPREPEDIDO'] + str(linha)].value != None and aba_ativa[dicioConfig['PEDIDO'] + str(linha)].value == None:
                                
                                prePedidoPendente = aba_ativa[dicioConfig['NUMPREPEDIDO']+str(linha)].value
                                
                                page.goto(f'https://www.me.com.br/VerPrePedidoWF.asp?Pedido={prePedidoPendente}&SuperCleanPage=false&Origin=home')
                                
                                statusPrePedido = page.locator('xpath=/html/body/main/div/div[1]/div[2]/div[2]/p[1]/span[2]').inner_html().split(' ')[0].strip()
                                
                                if statusPrePedido == 'APROVADO':
                                    numPedidoSAP = page.locator('xpath=/html/body/main/div/div[1]/div[1]/p[1]').inner_html().strip()
                                    aba_ativa[dicioConfig['PEDIDO']+str(linha)] = numPedidoSAP
                                    toaster.show_toast(f'Pré-Pedido {prePedidoPendente} aprovado!',f'O número do seu pedido é {numPedidoSAP}',icon_path='iconeME.ico', duration=8, threaded=True)
                                elif statusPrePedido != 'APROVADO':
                                    dataPrepedido = int(str(dicioConfig['TEMPO']).split('/')[0])
                                    hoje = int(date.today().strftime('%d/%m/%Y').split('/')[0])
                                    if hoje - dataPrepedido >= 2:
                                        toaster.show_toast(f'Pré-Pedido {prePedidoPendente} pendente há dois dias ou mais!',f'Solicite aprovação do resposável!',icon_path='iconeME.ico', duration=8, threaded=True)

                    try:
                        tabela.save(nomeplanilha)
                    except PermissionError:

                        toaster.show_toast(f'Erro ao salvar a planilha!',f'Provavelmente a planilha {nomeplanilha} está aberta no computador \nSalvando uma planilha temporaria no local!',icon_path='iconeME.ico', duration=10, threaded=True)
                        tabela.save(f'{nomeplanilhaSemext[0]}_Temp.{nomeplanilhaSemext[1]}')
                        
                except TimeoutError:
                    toaster.show_toast(f'Erro no monitoramento de requisição.',f'Não foi possivel monitorar as requisições. \nLentidão no mercado eletronico, tentando novamente em alguns minutos!',icon_path='iconeME.ico', duration=20, threaded=True)    
                    tabela.save(nomeplanilha)
                sleep(int(dicioConfig['TEMPO'])*60)
            elif quantPendente == 0:
               
                toaster.show_toast(f'Você não possui requisições pendentes!','Monitoramento desativado!',icon_path='iconeME.ico', duration=8, threaded=True)
                monitorReq.deselect()
            else:
                continue
        else:
            toaster.show_toast(f'Monitoramento desativado!',f'Você desativou o monitoramento de requisições.',icon_path='iconeME.ico', duration=8, threaded=True)
            break
      
def limpar():
    input_comentario.delete(0,"end")
    input_data.delete(0,"end")
    input_quantidade.delete(0,"end")
    input_titulo.delete(0,"end")
    input_valorUN.delete(0,"end")
    input_arquivo.delete("1.0","end")
    combo_categoria.set("")
    combo_centroCusto.set("")
    combo_filial.set("")
    combo_item.set("")
    combo_tipo.set("")
    checkNavegador.deselect()
    progress_bar.place_forget()
    caixa_numero_req.place_forget()
    caixa_numero_req.delete("1.0", 'end')
    mensagem_titulo['text'] = ('')
    mensagem_numero_req['text'] = ('')
    titulo_progress_bar['text'] = ('')
    habilitaProcurar()
    
def procurarArquivos(janela):
    global filenames
    filenames = askopenfilenames(
        title='Procurar arquivos',
    )
    
    if janela == 'configExcel':
        entryNomeArquivo.delete(0, "end")
        entryNomeArquivo.insert(INSERT, filenames[0])
    else:
        input_arquivo.insert(INSERT, filenames[0])
    
def habilitaProcurar(*args):
    if combo_categoria.get() == "PEDIDO REGULARIZACAO":
        arquivo_requisicao.place(relx=.33, rely=.5)
        input_arquivo.place(relx=.02, rely=.54)
        botaoProcuraArquivo.place(relx=.72, rely=.53)
        comentario_requisicao.place(relx=.41, rely=.6)
        input_comentario.place(relx=.15, rely=.64)
        botaoCriar.place(relx=.02, rely=.69)
        botaoCancelar.place(relx=.42, rely=.69)
        botaoLimpar.place(relx=.83, rely=.69)
        titulo_progress_bar.place(relx= .1, rely=.75)
        
        
    else:
        arquivo_requisicao.place_forget()
        input_arquivo.place_forget()
        botaoProcuraArquivo.place_forget()
        comentario_requisicao.place(relx=.41, rely=.5)
        input_comentario.place(relx=.15, rely=.54)
        botaoCriar.place(relx=.02, rely=.60)
        botaoCancelar.place(relx=.42, rely=.60)
        botaoLimpar.place(relx=.83, rely=.60)
        titulo_progress_bar.place(relx= .1, rely=.68)

def atualizaCodigo(*args):
    codLista.append(dicioTipo[combo_tipo.get()])
    combo_item.set(';'.join(codLista).strip())

def criarRequisicao(*args):
    tabela = load_workbook(dicioConfig['NOMEARQUIVO'], data_only=True)
    aba_ativa = tabela[dicioConfig['NOMEABA']]
    ultimaLinha = dicioConfig['REQUISICAO'] + str(len(aba_ativa[dicioConfig['REQUISICAO']])+1)
    comentario = input_comentario.get()
    caminho_arquivo = list(filenames)
    centro_custo = combo_centroCusto.get()
    cat_Pedido = combo_categoria.get()
    titulo_requisicao = input_titulo.get()
    item = combo_item.get().strip().split(";")
    valorun = input_valorUN.get().strip().split(";")
    quant = input_quantidade.get().split(";")
    data_esperada = input_data.get_date().strftime("%d/%m/%Y")
    filial = combo_filial.get()
    nome_filial = filial.split('-',1)[1][1:]
    
    progress_bar.place_forget()
    caixa_numero_req.place_forget()
    caixa_numero_req.delete("1.0", 'end')
    mensagem_titulo['text'] = ('')
    mensagem_numero_req['text'] = ('')
    titulo_progress_bar['text'] = ('')
    
    
    try:
        with sync_playwright() as p:
            if combo_categoria.get() == "PEDIDO REGULARIZACAO":
                progress_bar.place(relx= .10, rely=.79)
            else:
                progress_bar.place(relx= .10, rely=.72)
            # valorTotal = str(float(valorun) * int(quant))
            if varcheckNavegador.get():
                browser = p.chromium.launch(channel="chrome",headless=False)
            else:
                browser = p.chromium.launch(channel="chrome")
            page = browser.new_page()
            page.goto(dicioLogin['site'])

            # LOGIN ME
            titulo_progress_bar['text'] = ('Efetuando login no ME')      
            progress_bar['value'] += 14.28
            page.locator('xpath=//*[@id="LoginName"]').fill(dicioLogin['usuario_me'])
            page.locator('xpath=//*[@id="RAWSenha"]').fill(dicioLogin['senha_me'])
            page.locator('xpath=//*[@id="SubmitAuth"]').click()

            # CONFIGURAÇÃO DA REQUISIÇÃO
            titulo_progress_bar['text'] = ('Configurando a requisição')      
            progress_bar['value'] += 14.28
            page.locator('xpath=//*[@id="__layout"]/div/main/div/div/div[2]/div/div[1]/section/div/div/div/button').click()
            page.locator('xpath=//*[@id="__layout"]/div/main/div/div/div[2]/div/div[1]/section/div[1]/div[2]/ul/li[1]/a').click()
            page.locator('xpath=//*[@id="__layout"]/div/main/div/div/div[2]/div/div[1]/section/div[1]/div[2]/ul[2]/li[1]/a').click()
            frame = page.frame_locator('#PopUpConfiguration-if')
            page.wait_for_timeout(1000)
            frame.locator('xpath=//*[@id="select2-Categoria_Value-container"]').click()
            page.wait_for_timeout(500)
            frame.locator('xpath=/html/body/span/span/span[1]/input').fill(cat_Pedido)
            page.wait_for_timeout(500)
            frame.locator('xpath=/html/body/span/span/span[1]/input').press('Enter')
            page.wait_for_timeout(500)
            frame.locator('xpath=//*[@id="BOrgs_1__BorgDescription"]').press('Tab')
            page.wait_for_timeout(500)
            if frame.locator('xpath=//*[@id="BOrgs_0__BorgDescription"]').input_value().strip() == "":
                frame.locator('xpath=//*[@id="BOrgs_0__BorgDescription"]').fill("1 - VERO S.A.")
            frame.locator('xpath=//*[@id="BOrgs_1__BorgDescription"]').press('Tab')
            page.wait_for_timeout(500)
            frame.locator('xpath=//*[@id="BOrgs_1__BorgDescription"]').fill(filial)
            frame.locator('xpath=//*[@id="BOrgs_1__BorgDescription"]').press('Tab')
            frame.locator('xpath=//*[@id="btnSave"]').click()

            # SELECIONA ITENS E QUANTIDADES
            titulo_progress_bar['text'] = ('Adicionando itens e quantidades')      
            progress_bar['value'] += 14.28
            for i in range(len(quant)):
                page.wait_for_timeout(1000)
                page.locator('xpath=//*[@id="Valor"]').fill(item[i])
                page.locator('xpath=//*[@id="btnSearchSimple"]').click()
                page.locator('.icon-shopping-cart').click()
                page.wait_for_timeout(1000)
                page.keyboard.press('Control+A')
                page.wait_for_timeout(1000)
                page.keyboard.type(quant[i])
                page.wait_for_timeout(1000)
                page.keyboard.press('Tab')
                page.wait_for_timeout(1000)
            page.locator('xpath=//*[@id="btnAvancar"]').click()

            # TELA CONDIÇÕES GERAIS
            titulo_progress_bar['text'] = ('Ajustando condições gerais')      
            progress_bar['value'] += 14.28
            page.locator('xpath=//*[@id="Titulo_Value"]').fill(titulo_requisicao)
            page.locator('xpath=//*[@id="DataEsperada_Value"]').fill(data_esperada)
            page.wait_for_timeout(500)
            page.locator('xpath=//*[@id="select2-LocalEntrega_Value-container"]').click()
            page.locator('xpath=/html/body/span/span/span[1]/input').fill(nome_filial)
            page.locator('xpath=/html/body/span/span/span[1]/input').press('Enter')
            page.locator('xpath=//*[@id="CentroCusto_Text"]').fill(centro_custo[:4])
            page.wait_for_timeout(1000)
            page.locator('xpath=//*[@id="ui-id-2"]').click()
            page.locator('xpath=//*[@id="select2-LocalFaturamento_Value-container"]').click()
            page.locator('xpath=/html/body/span/span/span[1]/input').fill(nome_filial)
            page.locator('xpath=/html/body/span/span/span[1]/input').press('Enter')
            page.locator('xpath=//*[@id="Observacao_Value"]').fill(comentario)
            page.locator('xpath=//*[@id="btnAvancar"]').click()
            
            #            TELA DETALHES DOS ITENS
            titulo_progress_bar['text'] = ('Ajustando detalhes dos itens')      
            progress_bar['value'] += 14.28

            for i in range(len(quant)):
                valorun[i] = valorun[i].replace(',','.')
                valorTotal = str(float(valorun[i]) * int(quant[i]))
                page.locator(f'xpath=//*[@id="Itens_{i}__PrecoEstimado_Value"]').fill(valorun[i].replace(".",","))
                page.locator(f'xpath=//*[@id="select2-Itens_{i}__CategoriaContabil_Value-container"]').click()
                page.locator(f'xpath=//*[@id="select2-Itens_{i}__CategoriaContabil_Value-container"]').press('Enter')
                page.wait_for_timeout(1000)
                if cat_Pedido == 'PEDIDO REGULARIZACAO':
                    page.locator(f'xpath=//*[@id="Itens_{i}__Attributes_0__valor"]').fill(valorTotal.replace(".",","))
                else:
                    page.locator(f'xpath=//*[@id="Itens_{i}__Attributes_1__valor"]').fill(titulo_requisicao)
                    page.locator(f'xpath=//*[@id="Itens_{i}__Attributes_0__valor"]').fill(valorTotal.replace(".",","))
            page.locator('xpath=//*[@id="btnAvancar"]').click()
            page.wait_for_timeout(1000)


            # FINALIZAR REQUISIÇÃO
            titulo_progress_bar['text'] = ('Finalizando a requisição')    
             
            progress_bar['value'] += 14.28
            
            if cat_Pedido == 'PEDIDO REGULARIZACAO':
                with page.expect_popup() as popup_info:
                    page.locator('xpath=//*[@id="anexoReq_link"]').click()
                    popup = popup_info.value
                    popup.wait_for_load_state()
                    popup.locator('xpath=//*[@id="fuArquivo"]').set_input_files(caminho_arquivo)
                    popup.locator('xpath=//*[@id="ctl00_conteudo_formUpload_btn_ctl00_conteudo_formUpload_btnEnviar"]').click()
                    popup.wait_for_load_state()
                    popup.close()

            page.locator('xpath=//*[@id="btnAvancar"]').click()
            requisicao = page.locator('.badge-code ')
            page.wait_for_timeout(1000)
            progress_bar['value'] += 14.3

            if cat_Pedido == 'PEDIDO REGULARIZACAO':
                aba_ativa[ultimaLinha] = requisicao.inner_html().strip()[4:]

            if combo_categoria.get() == "PEDIDO REGULARIZACAO":
                caixa_numero_req.place(relx= .60, rely=.925)
                mensagem_titulo.place(relx= .1, rely=.85)
                mensagem_numero_req.place(relx= .1, rely=.92)
                
            else:
                caixa_numero_req.place(relx= .64, rely=.825)
                mensagem_titulo.place(relx= .1, rely=.77)
                mensagem_numero_req.place(relx= .1, rely=.82)

            titulo_progress_bar['text'] = ('######## REQUISIÇÃO FINALIZADA ########')
            caixa_numero_req.insert(INSERT, requisicao.inner_html().strip()[4:])
            mensagem_titulo['text'] = (titulo_requisicao)
            mensagem_numero_req['text'] = ('Sua requisição é: ')
            aba_ativa[dicioConfig['DATAABERTURA'] + str(linha)] = date.today().strftime('%d/%m/%Y')

    except TimeoutError:
        toaster.show_toast(f'Erro ao criar a requisição.',f'Lentidão no mercado eletronico!\nTente novamente em alguns minutos!',icon_path='iconeME.ico', duration=20, threaded=True)    
      
def salvarArquivo(nome_arquivo):

    if nome_arquivo == 'credenciais.txt':
        dicioLogin["usuario_me"] = credencialUser.get()
        dicioLogin["senha_me"] = credencialSenha.get()
        dicioLogin["site"] = credencialSite.get()
        with open(nome_arquivo, 'w') as credenciais:
            credencial=f"usuario_me={credencialUser.get()}\nsenha_me={credencialSenha.get()}\nsite={credencialSite.get()}"
            credenciais.write(credencial)
        janelaArquivos.destroy()

    elif nome_arquivo == 'filiais.txt':
        with open(nome_arquivo, 'w') as filiais:
            conteudo = list(listbox.get(0,"end"))
            for item in conteudo:
                filiais.write(item+'\n')
        janelaArquivos.destroy()

    elif nome_arquivo == 'codigos.txt':
        with open(nome_arquivo, 'w') as codigos:
            conteudo = list(listbox.get(0,"end"))
            for item in conteudo:
                codigos.write(item+'\n')
        janelaArquivos.destroy()

    elif nome_arquivo == 'categorias.txt':
        with open(nome_arquivo, 'w') as categorias:
            conteudo = list(listbox.get(0,"end"))
            for item in conteudo:
                categorias.write(item+'\n')
        janelaArquivos.destroy()

    elif nome_arquivo == 'centrocustos.txt':
        with open(nome_arquivo, 'w') as categorias:
            conteudo = list(listbox.get(0,"end"))
            for item in conteudo:
                categorias.write(item+'\n')
        janelaArquivos.destroy()

    elif nome_arquivo == 'TipoRequisicao.txt':
        with open(nome_arquivo, 'w') as tiporeq:

            for k, v in dicioTipo.items():
                tiporeq.write(k+';'+v+'\n')
        janelaArquivos.destroy()

    elif nome_arquivo == 'configExcel.txt':
        with open(nome_arquivo, 'w') as configE:
            dicioConfig['NF'] = EntryNF.get().upper()
            dicioConfig['REQUISICAO'] =  entryRequisicao.get().upper()
            dicioConfig['DATAABERTURA'] =  entryDataAbertura.get().upper()
            dicioConfig['CNPJ'] =  entryCNPJ.get().upper()
            dicioConfig['NUMPREPEDIDO'] =  entryNumPrePedido.get().upper()
            dicioConfig['DATAPREPEDIDO'] =  entryDataPrePedido.get().upper()
            dicioConfig['PEDIDO'] =  entryPedido.get().upper()
            dicioConfig['STATUS'] =  entryStatus.get().upper()
            dicioConfig['NOMEABA'] =  entryNomeAba.get().upper()
            dicioConfig['NOMEARQUIVO'] =  entryNomeArquivo.get()
            dicioConfig['TEMPO'] =  entryTempoMonitor.get()
            for k, v in dicioConfig.items():
                configE.write(k+';'+v+'\n')
        janelaConfig.destroy()
   
def addListBox(nome_arquivo):
    if nome_arquivo == 'TipoRequisicao.txt':
        dicioTipo[caixa_adicionar.get()] = caixa_item.get()
        listbox.insert("end", caixa_adicionar.get())
    
    else:
        listbox.insert("end", caixa_adicionar.get())

def remListBox(nome_arquivo):
    if nome_arquivo == 'TipoRequisicao.txt':
        dicioTipo.pop(listbox.get(listbox.curselection()))
        listbox.delete(listbox.curselection())

    else:
        listbox.delete(listbox.curselection())

def abreArquivo(nome_arquivo):
    global credencialUser
    global credencialSite
    global credencialSenha
    global listbox
    global janelaArquivos
    global caixa_item
    global caixa_adicionar

    credencialUser = StringVar()
    credencialSite = StringVar()
    credencialSenha = StringVar()
    caixa_item = StringVar()

    #Cria segunda janela
    titulo = nome_arquivo.split('.')[0]
    janelaArquivos = Toplevel(janela)
    janelaArquivos.title(f'Editar informações de {titulo}')
    mensagem = Label(janelaArquivos, text='', font='calibri 11')
    mensagemTitulo = Label(janelaArquivos, text='', font='calibri 11')
    labelSelecaoTitulo = Label(janelaArquivos, text='', font='calibri 11')
    labelSelecao = Label(janelaArquivos, text='', font='calibri 11')
    listbox = Listbox(janelaArquivos, height=13, width=30)
    caixa_item = Entry(janelaArquivos, textvariable=caixa_item)
    caixa_adicionar = Entry(janelaArquivos, textvariable=credencialUser, width=43)
    
    labelUser = Label(janelaArquivos,text='Usuário:')
    labelSenha = Label(janelaArquivos,text='Senha:')
    labelSite = Label(janelaArquivos,text='Site:')
    
    credencialUser = Entry(janelaArquivos, textvariable=credencialUser, width=20)
    credencialSenha = Entry(janelaArquivos, textvariable=credencialSenha, width=20)
    credencialSite = Entry(janelaArquivos, textvariable= credencialSite, width=20)
    
    botao_salvar = Button(janelaArquivos, text='Salvar', width=12, command=lambda: salvarArquivo(nome_arquivo))
    botao_cancelar = Button(janelaArquivos, text='Cancelar', width=12, command=janelaArquivos.destroy)
    botao_adicionar = Button(janelaArquivos, text='Adicionar', width=12, command= lambda: addListBox(nome_arquivo))
    botao_remover = Button(janelaArquivos, text='Remover', width=12, command=lambda: remListBox(nome_arquivo))
    
    #Abre arquivo
    arquivo = open(nome_arquivo, "r", encoding="ISO-8859-1")
    conteudo = arquivo.read().strip()

    if nome_arquivo == 'credenciais.txt':
        janelaArquivos.geometry('400x200')
        mensagem.configure(text='Adicionar informações de login!')
        mensagem.place(relx=.28, rely=0.02)
        labelUser.place(relx=0.25, rely=0.14)       
        labelSenha.place(relx=0.25, rely=0.29)       
        labelSite.place(relx=0.25, rely=0.44)       
        credencialUser.place(relx=.48, rely=0.15)  
        credencialSenha.place(relx=.48, rely=0.30)         
        credencialSite.place(relx=.48, rely=0.45)         
        credencialUser.insert(END, dicioLogin['usuario_me'])
        credencialSenha.insert(END, dicioLogin['senha_me'])
        credencialSite.insert(END, dicioLogin['site'])
        botao_salvar.place(relx=0.2, rely=0.83)
        botao_cancelar.place(relx=0.55, rely=0.83)
    
    elif nome_arquivo == 'filiais.txt':
        janelaArquivos.geometry('500x300')

        mensagemTitulo.place(relx=0.43, rely=0.08)
        mensagemTitulo.configure(text='Inserir filial:')
        caixa_adicionar.place(relx=0.43, rely=0.15)
        mensagem.configure(text='Controle de filiais.')
        mensagem.place(relx=0.1, rely=0.02)

        labelSelecaoTitulo.place(relx=0.43, rely=0.33)
        labelSelecaoTitulo.configure(text='Filial selecionada: ')
        labelSelecao.place(relx=0.43, rely=0.4)
        labelSelecao.configure(text='')

        var = Variable(value=conteudo.split('\n'))
        listbox.configure(listvariable=var)
        listbox.place(relx=.05, rely=0.15) 
        listbox.bind('<<ListboxSelect>>', lambda e: labelSelecao.configure(text=listbox.get(listbox.curselection())))
        botao_adicionar.place(relx=0.77, rely=0.23)
        botao_salvar.place(relx=0.1, rely=0.89)
        botao_cancelar.place(relx=0.77, rely=0.89)
        botao_remover.place(relx=0.77, rely=0.48)
 
    elif nome_arquivo == 'centrocustos.txt':
        janelaArquivos.geometry('500x300')

        mensagemTitulo.place(relx=0.43, rely=0.08)
        mensagemTitulo.configure(text='Inserir centro de custo:')
        caixa_adicionar.place(relx=0.43, rely=0.15)
        mensagem.configure(text='Controle de centro de custs.')
        mensagem.place(relx=0.1, rely=0.02)

        labelSelecaoTitulo.place(relx=0.43, rely=0.33)
        labelSelecaoTitulo.configure(text='Centro de custo selecionado: ')
        labelSelecao.place(relx=0.43, rely=0.4)
        labelSelecao.configure(text='')

        var = Variable(value=conteudo.split('\n'))
        listbox.configure(listvariable=var)
        listbox.place(relx=.05, rely=0.15) 
        listbox.bind('<<ListboxSelect>>', lambda e: labelSelecao.configure(text=listbox.get(listbox.curselection())))
        botao_adicionar.place(relx=0.77, rely=0.23)
        botao_salvar.place(relx=0.1, rely=0.89)
        botao_cancelar.place(relx=0.77, rely=0.89)
        botao_remover.place(relx=0.77, rely=0.48)

    elif nome_arquivo == 'categorias.txt':
        janelaArquivos.geometry('450x300')

        mensagemTitulo.place(relx=0.47, rely=0.08)
        mensagemTitulo.configure(text='Inserir categoria:')
        caixa_adicionar.place(relx=0.47, rely=0.15)
        caixa_adicionar.configure(width=35)
        mensagem.configure(text='Controle de categorias.')
        mensagem.place(relx=0.1, rely=0.02)

        labelSelecaoTitulo.place(relx=0.47, rely=0.33)
        labelSelecaoTitulo.configure(text='Categoria selecionada: ')
        labelSelecao.place(relx=0.47, rely=0.4)
        labelSelecao.configure(text='')

        var = Variable(value=conteudo.split('\n'))
        listbox.configure(listvariable=var)
        listbox.place(relx=.05, rely=0.15) 
        listbox.bind('<<ListboxSelect>>', lambda e: labelSelecao.configure(text=listbox.get(listbox.curselection())))
        botao_adicionar.place(relx=0.77, rely=0.23)
        botao_salvar.place(relx=0.1, rely=0.89)
        botao_cancelar.place(relx=0.77, rely=0.89)
        botao_remover.place(relx=0.77, rely=0.48)

    elif nome_arquivo == 'codigos.txt':
        janelaArquivos.geometry('300x300')
        labelSelecaoTitulo.place(relx=0.5, rely=0.36)
        labelSelecaoTitulo.configure(text='Item selecionado: ')
        labelSelecao.place(relx=0.5, rely=0.43)
        labelSelecao.configure(text='')
        var = Variable(value=conteudo.split('\n'))
        listbox.configure(listvariable=var, width=20)
        listbox.place(relx=.05, rely=0.15) 
        listbox.bind('<<ListboxSelect>>', lambda e: labelSelecao.configure(text=listbox.get(listbox.curselection())))
        mensagemTitulo.place(relx=0.5, rely=0.14)
        mensagemTitulo.configure(text='Inserir item:')
        caixa_adicionar.configure(width=20)
        caixa_adicionar.place(relx=0.5, rely=0.21)
        mensagem.configure(text='Controle de itens.')
        mensagem.place(relx=0.1, rely=0.02)
        botao_adicionar.place(relx=0.65, rely=0.28)
        botao_remover.place(relx=0.65, rely=0.53)
        botao_salvar.place(relx=0.1, rely=0.89)
        botao_cancelar.place(relx=0.65, rely=0.89)

    elif nome_arquivo == 'TipoRequisicao.txt':
        janelaArquivos.geometry('500x300')

        mensagem.configure(text='Controle de tipos de requisições.')
        mensagem.place(relx=0.05, rely=0.018)
        mensagemTitulo.place(relx=0.43, rely=0.08)
        mensagemTitulo.configure(text='Inserir código e descrição do item:')
        caixa_adicionar.place(relx=0.63, rely=0.15)
        caixa_adicionar.configure(width=25)
        caixa_item.place(relx=0.43, rely=0.15)
        caixa_item.configure(width=10)
        labelSelecaoTitulo.place(relx=0.43, rely=0.33)
        labelSelecaoTitulo.configure(text='Descrição selecionada: ')
        labelSelecao.place(relx=0.43, rely=0.4)
        labelSelecao.configure(text='')
        var = Variable(value=list(dicioTipo.keys()))
        listbox.configure(listvariable=var)
        listbox.place(relx=.05, rely=0.15) 
        listbox.bind('<<ListboxSelect>>', lambda e: labelSelecao.configure(text=dicioTipo[listbox.get(listbox.curselection())].strip() + " - " + listbox.get(listbox.curselection())))
        botao_adicionar.place(relx=0.77, rely=0.23)
        botao_salvar.place(relx=0.1, rely=0.89)
        botao_cancelar.place(relx=0.77, rely=0.89)
        botao_remover.place(relx=0.77, rely=0.48)

def threading():
    thread = Thread(target=monitorME)
    thread.start()
    toaster.show_toast(f'Monitoramento ativo!',f'O sistema irá monitorar as requisições a cada {tempoMonitor} minuto (s).',icon_path='iconeME.ico', duration=10, threaded=True)
            
def fechar():
    varmonitorReq = False
    try:
        janela.quit()
        janela.destroy()
        subprocess.call("TASKKILL /F /IM python.exe", shell=True)
        subprocess.call("TASKKILL /F /IM automacao.exe", shell=True)
    except:
        janela.quit()
        janela.destroy()

varcheckNavegador = BooleanVar()
varmonitorReq = BooleanVar()

checkNavegador = Checkbutton(janela, text='Abre Nav',variable=varcheckNavegador, offvalue=False, onvalue=True)
checkNavegador.place(relx=.02, rely=.02)

monitorReq = Checkbutton(janela, text='Monitor',variable=varmonitorReq, onvalue=True, offvalue=False, command=threading)
monitorReq.place(relx=.78, rely=.02)

titulo_requisicao = Label(janela, text='Titulo da requisição', font='calibri, 10')
titulo_requisicao.place(relx=.35, rely=.02)
input_titulo = Entry(janela, text='Titulo',width=55)
input_titulo.place(relx=.02, rely=.06)

tipo_requisicao = Label(janela, text='Tipo de requisição', font='calibri, 10')
tipo_requisicao.place(relx=.35, rely=.09)
combo_tipo = Combobox(janela, width=52, state="readonly")
combo_tipo['values']=(listaTipo)
combo_tipo.place(relx=.02, rely=.13)
combo_tipo.bind("<<ComboboxSelected>>", atualizaCodigo)

item_requisicao = Label(janela, text='Item', font='calibri, 10')
item_requisicao.place(relx=.02, rely=.17)
combo_item = Combobox(janela, width=20)
combo_item['values']=(codigos)
combo_item.place(relx=.02, rely=.21)

valorUN_requisicao = Label(janela, text='Valor unitário', font='calibri, 10')
valorUN_requisicao.place(relx=.57, rely=.17)
input_valorUN = Entry(janela, text='valorunitario')
input_valorUN.place(relx=.57, rely=.21)

quantidade_requisicao = Label(janela, text='Quantidade', font='calibri, 10')
quantidade_requisicao.place(relx=.02, rely=.25)
input_quantidade = Entry(janela, text='quantidade', width=23)
input_quantidade.place(relx=.02, rely=.29)

data_requisicao = Label(janela, text='Data esperada', font='calibri, 10')
data_requisicao.place(relx=.57, rely=.25)
input_data = DateEntry(janela, width=20)
input_data.delete(0,"end")
input_data.place(relx=.57, rely=.29)

categoria_requisicao = Label(janela, text='Selecione a categoria', font='calibri, 10')
categoria_requisicao.place(relx=.02, rely=.33)
combo_categoria = Combobox(janela, width=25, state="readonly")
combo_categoria['values']=(categorias)
combo_categoria.place(relx=.02, rely=.37)
combo_categoria.bind("<<ComboboxSelected>>", habilitaProcurar)

centroCusto_requisicao = Label(janela, text='Centro de custo', font='calibri, 10')
centroCusto_requisicao.place(relx=.57, rely=.33)
combo_centroCusto = Combobox(janela, width=20, state="readonly")
combo_centroCusto['values']=(centroCustos)
combo_centroCusto.place(relx=.57, rely=.37)

filial_requisicao = Label(janela, text='Selecione a filial:', font='calibri, 10')
filial_requisicao.place(relx=.34, rely=.41)
combo_filial = Combobox(janela, width=52, state="readonly")
combo_filial['values']=(filiais)
combo_filial.place(relx=.02, rely=.45)

arquivo_requisicao = Label(janela, text='Selecionar arquivos:', font='calibri, 10')
input_arquivo = Text(janela, width=28, height=1)
botaoProcuraArquivo = Button(janela, text='Procurar', command=lambda: procurarArquivos('PEDIDOREGULARIZACAO'), width=10, font='calibri, 10')

comentario_requisicao = Label(janela, text='Comentario:', font='calibri, 10')
comentario_requisicao.place(relx=.41, rely=.5)
input_comentario = Entry(janela, text='comentario', width=40)
input_comentario.place(relx=.15, rely=.54)

botaoCriar = Button(janela, text='Criar', command=lambda: Thread(target=criarRequisicao).start(), width=15, font='calibri, 10')
botaoCriar.place(relx=.02, rely=.60)
botaoCancelar = Button(janela, text='Cancelar',  width=15, font='calibri, 10', command=fechar)
botaoCancelar.place(relx=.42, rely=.60)
botaoLimpar = Button(janela, text='Limpar', width=5, font='calibri, 10', command=limpar)
botaoLimpar.place(relx=.83, rely=.60)

titulo_progress_bar = Label(janela, text="", font='calibri, 10')
titulo_progress_bar.place(relx= .1, rely=.68)
progress_bar = Progressbar(janela, orient= 'horizontal', mode='determinate', length=280)


mensagem_titulo = Label(janela, text="", font='calibri, 11')
mensagem_titulo.place(relx= .15, rely=.77)
mensagem_numero_req = Label(janela, text="", font='calibri, 11')
mensagem_numero_req.place(relx= .15, rely=.84)
caixa_numero_req = Text(janela, font='calibri, 10', width=15, height= 1)



janela.protocol("WM_DELETE_WINDOW", fechar)
janela.config(menu=menubar)

janela.mainloop()
varmonitorReq = False